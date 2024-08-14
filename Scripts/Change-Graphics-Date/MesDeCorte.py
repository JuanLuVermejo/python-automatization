import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def borrar_rango_celdas(archivo_excel):
    # Cargar el libro de trabajo
    wb = load_workbook(archivo_excel)

    # Verificar si la hoja 'BDDetalle' existe en el libro de trabajo
    if 'BDDetalle' in wb.sheetnames:
        # Seleccionar la hoja 'BDDetalle'
        ws = wb['BDDetalle']

        # Borrar el rango de celdas A76:I76 queda hasta febrero (75)
        for row in ws.iter_rows(min_row=74, max_row=80, min_col=1, max_col=9):
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type='none')  # Eliminar formato de relleno

        # Guardar los cambios en el archivo Excel
        wb.save(archivo_excel)
        print(f"Se borró el rango de celdas y el formato en 'BDDetalle' del archivo '{archivo_excel}'")
    else:
        print(f"No se encontró la hoja 'BDDetalle' en el archivo '{archivo_excel}'")

def main():
    # Directorio donde se encuentran los archivos Excel
    directorio = r'C:\Users\juan.vermejo\Documents\CPNO\00. Informes por Analizar'

    # Recorrer todos los archivos en el directorio
    for filename in os.listdir(directorio):
        if filename.endswith('.xlsx'):
            archivo_excel = os.path.join(directorio, filename)
            borrar_rango_celdas(archivo_excel)

main()
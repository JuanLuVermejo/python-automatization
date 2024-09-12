import os
import pandas as pd
import openpyxl

def process_excel_file(file_path):
    try:
        # Cargar el archivo de Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Verificar si las hojas "BD" y "Hoja de Calculos" existen
        if "BD" not in wb.sheetnames or "Hoja de Calculos" not in wb.sheetnames:
            return None

        # Obtener el valor de la celda C3 en la hoja "BD"
        sheet_bd = wb['BD']
        c3_value = sheet_bd['C3'].value
        
        # Leer los valores de las celdas P2 y C2 en la hoja "Hoja de Calculos"
        sheet = wb['Hoja de Calculos']
        p2_value = sheet['P2'].value
        c2_value = sheet['C2'].value

        # Imprimir el valor de C2
        print(f'Procesando archivo: {file_path}')
        print(f'Cuenta Contrato (C2): {c2_value}')

        # Encontrar las filas con "Item" y "Total m3/h"
        item_row = None
        total_row = None

        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=2).value
            if cell_value == "Item":
                item_row = row
            elif cell_value == "Total m3/h":
                total_row = row
                break

        # Asegurarse de que se encontraron ambas filas
        if item_row is None or total_row is None:
            raise ValueError(f"No se encontraron las filas 'Item' y 'Total m3/h' en {file_path}.")

        # Leer los valores entre las filas encontradas
        data = []
        count_numerics = 0
        for row in range(item_row + 1, total_row):
            equipo = sheet.cell(row=row, column=3).value
            potencia = sheet.cell(row=row, column=4).value
            if isinstance(potencia, (int, float)):  # Verificar si el valor es numérico
                count_numerics += 1
            data.append((equipo, potencia))

        # Crear una fila para el DataFrame
        row_data = {
            'Archivo': file_path, 
            'Nombre del cliente': c3_value, 
            'Cuenta Contrato': c2_value, 
            'Tipo de medidor': p2_value,
            'Cantidad de equipos': count_numerics
        }

        for i, (equipo, potencia) in enumerate(data):
            row_data[f'Equipo {i + 1}'] = equipo
            row_data[f'Potencia {i + 1}'] = potencia

        return row_data
    
    except Exception as e:
        print(f"Error procesando {file_path}: {e}")
        return None

def process_folder(folder_path, output_file):
    # Crear una lista para almacenar los resultados
    all_data = []

    # Iterar sobre todos los archivos en la carpeta y subcarpetas
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith(".xlsx"):
                file_path = os.path.join(root, filename)
                row_data = process_excel_file(file_path)
                if row_data:
                    all_data.append(row_data)

    # Convertir la lista de resultados en un DataFrame
    if all_data:
        df = pd.DataFrame(all_data)
        # Guardar el DataFrame en un archivo Excel
        df.to_excel(output_file, index=False)
        print(f'Resultados guardados en {output_file}')
    else:
        print("No se encontraron archivos con las hojas necesarias.")

# Ejemplo de uso
folder_path = r'C:\Users\juan.vermejo\Documents\CPNO\Pruebas\Masivo\Comedores Populares'
# folder_path = r'C:\Users\juan.vermejo\Documents\CPNO\Pruebas\Masivo'
output_file = r'C:\Users\juan.vermejo\Documents\CPNO\Pruebas\Masivo\Resultados\Mst - Potencias Registradas (Comedores Populares).xlsx'
process_folder(folder_path, output_file)
